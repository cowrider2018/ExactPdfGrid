"""
Step 5: Resolve the logical grid positions of all cells, taking merged
        (multi-row / multi-column) cells into account.
Step 6: Write the result to an .xlsx file via openpyxl.

The CellRegion objects already carry:
  .row / .col          - 0-based grid indices
  .row_span / .col_span - how many logical rows/cols the cell occupies
  .text                 - extracted text

This module converts those to a properly merged openpyxl workbook.
"""

from __future__ import annotations
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .detection import CellRegion


# Fixed Excel styling constants (not user-tunable; they describe Excel concepts).
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")

# Default calibration: ~96 px = 1 inch = 8.43 Excel column-width units at default font.
DEFAULT_PX_PER_CHAR_CALIBRATION = 96 / 8.43
DEFAULT_MIN_COL_WIDTH = 4.0
DEFAULT_MIN_ROW_HEIGHT = 12.0


# ---------------------------------------------------------------------------
# Step 5 - sanity-check and sort the grid
# ---------------------------------------------------------------------------

def resolve_grid(cells: list[CellRegion]) -> list[CellRegion]:
    """
    Sort cells into reading order (top-to-bottom, left-to-right) and
    print a brief summary of the grid dimensions and merged cells.
    """
    if not cells:
        return []

    sorted_cells = sorted(cells, key=lambda c: (c.row, c.col))

    max_row = max(c.row + c.row_span for c in sorted_cells)
    max_col = max(c.col + c.col_span for c in sorted_cells)
    merged = [c for c in sorted_cells if c.row_span > 1 or c.col_span > 1]

    print(f"  [output] grid size: {max_row} rows x {max_col} cols")
    print(f"  [output] merged cells: {len(merged)}")
    for mc in merged:
        print(
            f"    r{mc.row}c{mc.col} spans {mc.row_span}x{mc.col_span} "
            f"-> text={mc.text!r}"
        )

    return sorted_cells


# ---------------------------------------------------------------------------
# Step 6 - write to xlsx
# ---------------------------------------------------------------------------

def _estimate_col_widths(
    xs: list[int],
    dpi: int = 200,
    min_col_width: float = DEFAULT_MIN_COL_WIDTH,
    px_per_char_calibration: float = DEFAULT_PX_PER_CHAR_CALIBRATION,
) -> list[float]:
    """Convert pixel column widths to approximate Excel column widths (characters)."""
    px_per_char = dpi / px_per_char_calibration
    widths = []
    for i in range(len(xs) - 1):
        px = xs[i + 1] - xs[i]
        widths.append(max(min_col_width, px / px_per_char))
    return widths


def _estimate_row_heights(
    ys: list[int],
    dpi: int = 200,
    min_row_height: float = DEFAULT_MIN_ROW_HEIGHT,
) -> list[float]:
    """Convert pixel row heights to Excel row heights (points)."""
    px_per_pt = dpi / 72.0
    heights = []
    for i in range(len(ys) - 1):
        px = ys[i + 1] - ys[i]
        heights.append(max(min_row_height, px / px_per_pt))
    return heights


def _populate_worksheet(
    ws,
    cells: list[CellRegion],
    ys: list[int],
    xs: list[int],
    dpi: int = 200,
    apply_borders: bool = True,
    apply_size_hints: bool = True,
    min_col_width: float = DEFAULT_MIN_COL_WIDTH,
    min_row_height: float = DEFAULT_MIN_ROW_HEIGHT,
    px_per_char_calibration: float = DEFAULT_PX_PER_CHAR_CALIBRATION,
) -> None:
    """Write a single worksheet with the given grid data."""
    if apply_size_hints and len(xs) >= 2 and len(ys) >= 2:
        col_widths = _estimate_col_widths(
            xs, dpi, min_col_width=min_col_width,
            px_per_char_calibration=px_per_char_calibration,
        )
        row_heights = _estimate_row_heights(
            ys, dpi, min_row_height=min_row_height,
        )

        for ci, cw in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(ci + 1)].width = cw
        for ri, rh in enumerate(row_heights):
            ws.row_dimensions[ri + 1].height = rh

    for cell in cells:
        xl_row = cell.row + 1
        xl_col = cell.col + 1
        ws_cell = ws.cell(row=xl_row, column=xl_col, value=cell.text or None)
        ws_cell.alignment = _WRAP
        if apply_borders:
            ws_cell.border = _BORDER

        if cell.row_span > 1 or cell.col_span > 1:
            merge_row_end = xl_row + cell.row_span - 1
            merge_col_end = xl_col + cell.col_span - 1
            ws.merge_cells(
                start_row=xl_row, start_column=xl_col,
                end_row=merge_row_end, end_column=merge_col_end,
            )
            ws_cell.alignment = _WRAP
            if apply_borders:
                ws_cell.border = _BORDER


def write_xlsx(
    cells: list[CellRegion],
    ys: list[int],
    xs: list[int],
    out_path: str,
    sheet_name: str = "Table",
    dpi: int = 200,
    apply_borders: bool = True,
    apply_size_hints: bool = True,
    min_col_width: float = DEFAULT_MIN_COL_WIDTH,
    min_row_height: float = DEFAULT_MIN_ROW_HEIGHT,
    px_per_char_calibration: float = DEFAULT_PX_PER_CHAR_CALIBRATION,
) -> None:
    """Write all cells to an xlsx workbook with a single sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    _populate_worksheet(
        ws, cells, ys, xs, dpi, apply_borders, apply_size_hints,
        min_col_width=min_col_width,
        min_row_height=min_row_height,
        px_per_char_calibration=px_per_char_calibration,
    )
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  [output] saved -> {out_path}")


def write_xlsx_workbook(
    sheets: list[tuple[str, list[CellRegion], list[int], list[int]]],
    out_path: str,
    dpi: int = 200,
    apply_borders: bool = True,
    apply_size_hints: bool = True,
    min_col_width: float = DEFAULT_MIN_COL_WIDTH,
    min_row_height: float = DEFAULT_MIN_ROW_HEIGHT,
    px_per_char_calibration: float = DEFAULT_PX_PER_CHAR_CALIBRATION,
) -> None:
    """Write multiple sheets into the same xlsx workbook."""
    if not sheets:
        print("  [output] no sheets to write.")
        return

    wb = openpyxl.Workbook()
    title, cells, ys, xs = sheets[0]
    ws = wb.active
    ws.title = title
    _populate_worksheet(
        ws, cells, ys, xs, dpi, apply_borders, apply_size_hints,
        min_col_width=min_col_width,
        min_row_height=min_row_height,
        px_per_char_calibration=px_per_char_calibration,
    )

    for title, cells, ys, xs in sheets[1:]:
        ws = wb.create_sheet(title=title)
        _populate_worksheet(
            ws, cells, ys, xs, dpi, apply_borders, apply_size_hints,
            min_col_width=min_col_width,
            min_row_height=min_row_height,
            px_per_char_calibration=px_per_char_calibration,
        )

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  [output] saved workbook with {len(sheets)} sheets -> {out_path}")
