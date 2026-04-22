"""
Step 5: Resolve the logical grid positions of all cells, taking merged
        (multi-row / multi-column) cells into account.
Step 6: Write the result to an .xlsx file via openpyxl.

The CellRegion objects already carry:
  .row / .col          – 0-based grid indices
  .row_span / .col_span – how many logical rows/cols the cell occupies
  .text                 – OCR result

This module converts those to a properly merged openpyxl workbook.
"""

from __future__ import annotations
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from table_detection import CellRegion


# ---------------------------------------------------------------------------
# Step 5 – sanity-check and sort the grid
# ---------------------------------------------------------------------------

def resolve_grid(cells: list[CellRegion]) -> list[CellRegion]:
    """
    Sort cells into reading order (top-to-bottom, left-to-right) and
    print a brief summary of the grid dimensions and merged cells.

    This is the place to apply any heuristic corrections if the
    table_detection step occasionally mis-classifies a span; currently
    we just validate and sort.

    Returns
    -------
    Sorted copy of the cell list.
    """
    if not cells:
        return []

    sorted_cells = sorted(cells, key=lambda c: (c.row, c.col))

    max_row = max(c.row + c.row_span for c in sorted_cells)
    max_col = max(c.col + c.col_span for c in sorted_cells)
    merged = [c for c in sorted_cells if c.row_span > 1 or c.col_span > 1]

    print(f"  [xlsx_output] grid size: {max_row} rows × {max_col} cols")
    print(f"  [xlsx_output] merged cells: {len(merged)}")
    for mc in merged:
        print(
            f"    r{mc.row}c{mc.col} spans {mc.row_span}×{mc.col_span} "
            f"→ text={mc.text!r}"
        )

    return sorted_cells


# ---------------------------------------------------------------------------
# Step 6 – write to xlsx
# ---------------------------------------------------------------------------

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")


def _estimate_col_widths(
    cells: list[CellRegion], xs: list[int], dpi: int = 200
) -> list[float]:
    """
    Convert pixel column widths to approximate Excel column widths (characters).
    Calibration: ~96 px ≈ 1 inch ≈ 8.43 Excel column-width units at default font.
    """
    px_per_char = dpi / (96 / 8.43)  # ≈ ~17.5 px per char-unit at 200 dpi
    widths = []
    for i in range(len(xs) - 1):
        px = xs[i + 1] - xs[i]
        widths.append(max(4.0, px / px_per_char))
    return widths


def _estimate_row_heights(
    cells: list[CellRegion], ys: list[int], dpi: int = 200
) -> list[float]:
    """
    Convert pixel row heights to Excel row heights (points).
    1 inch = 72 pt;  at 200 dpi, 1 pt = 200/72 ≈ 2.78 px.
    """
    px_per_pt = dpi / 72.0
    heights = []
    for i in range(len(ys) - 1):
        px = ys[i + 1] - ys[i]
        heights.append(max(12.0, px / px_per_pt))
    return heights


def write_xlsx(
    cells: list[CellRegion],
    ys: list[int],
    xs: list[int],
    out_path: str,
    sheet_name: str = "Table",
    dpi: int = 200,
    apply_borders: bool = True,
    apply_size_hints: bool = True,
) -> None:
    """
    Write all cells to an xlsx workbook.

    Parameters
    ----------
    cells       : resolved cell list (from resolve_grid)
    ys          : row boundary y-pixel positions (len = n_rows + 1)
    xs          : col boundary x-pixel positions (len = n_cols + 1)
    out_path    : destination .xlsx file path
    sheet_name  : worksheet tab name
    dpi         : DPI used when rendering PDF (for column-width estimation)
    apply_borders : draw thin borders around every cell
    apply_size_hints : set approximate row heights and col widths
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # ---- column widths and row heights -------------------------------------
    if apply_size_hints and len(xs) >= 2 and len(ys) >= 2:
        col_widths = _estimate_col_widths(cells, xs, dpi)
        row_heights = _estimate_row_heights(cells, ys, dpi)

        for ci, cw in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(ci + 1)].width = cw
        for ri, rh in enumerate(row_heights):
            ws.row_dimensions[ri + 1].height = rh

    # ---- write cells -------------------------------------------------------
    for cell in cells:
        # openpyxl is 1-based
        xl_row = cell.row + 1
        xl_col = cell.col + 1

        # Write text to the top-left cell of the span
        ws_cell = ws.cell(row=xl_row, column=xl_col, value=cell.text or None)
        ws_cell.alignment = _WRAP
        if apply_borders:
            ws_cell.border = _BORDER

        # Merge if span > 1
        if cell.row_span > 1 or cell.col_span > 1:
            merge_row_end = xl_row + cell.row_span - 1
            merge_col_end = xl_col + cell.col_span - 1
            ws.merge_cells(
                start_row=xl_row, start_column=xl_col,
                end_row=merge_row_end, end_column=merge_col_end,
            )
            # openpyxl requires the top-left cell to carry style after merge
            ws_cell.alignment = _WRAP
            if apply_borders:
                ws_cell.border = _BORDER

    # ---- save --------------------------------------------------------------
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  [xlsx_output] saved → {out_path}")
